#!/usr/bin/env python3
"""
BKS Studio — Price Audit & Update Tool
=======================================
Versione: 1.0 — 16 Giugno 2026
Autore: Gaetano (Claude Sonnet 4.6) per Roberto Picchioni

FUNZIONI:
  1. Legge il CSV Shopify con tutti i prodotti e prezzi attuali
  2. Carica la tabella costi Printify (manuale o da file)
  3. Calcola margine reale per ogni prodotto
  4. Genera report semaforo: 🟢 OK / 🟡 ALERT / 🔴 STOP
  5. Propone il prezzo corretto per ogni prodotto fuori range
  6. Genera CSV di correzione pronto per import Shopify
  7. (Opzionale) Aggiorna i prezzi via Shopify Admin API

USO:
  pip install pandas requests python-dotenv openpyxl
  python3 bks_price_audit.py --csv shopify_export.csv
  python3 bks_price_audit.py --csv shopify_export.csv --update  # aggiorna via API

INPUT:
  - shopify_export.csv: export da Shopify Admin > Products > Export
  - costs_table.csv: tabella costi Printify (opzionale, usa default se assente)
  - .env: SHOPIFY_STORE, SHOPIFY_TOKEN (solo per --update)

OUTPUT:
  - bks_price_audit_report.xlsx: report completo con semaforo
  - bks_price_corrections.csv: CSV pronto per import Shopify
  - bks_price_corrections_api.json: payload per aggiornamento API
"""

import os
import sys
import json
import argparse
import pandas as pd
import requests
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─────────────────────────────────────────────────────────────────────────────
# 1. TABELLA COSTI PRINTIFY (valori medi — da aggiornare con dati reali)
# ─────────────────────────────────────────────────────────────────────────────
# Struttura: category_keyword → {base_cost_min, base_cost_max, base_cost_avg,
#                                 shipping_eu_min, shipping_eu_max, shipping_eu_avg,
#                                 margin_target, price_min, price_target, price_max}

COST_TABLE = {
    "Sneaker": {
        "base_min": 18.0, "base_max": 26.0, "base_avg": 22.0,
        "ship_min": 8.0,  "ship_max": 14.0, "ship_avg": 11.0,
        "margin_target": 0.60,
        "price_min": 75.0, "price_target": 89.0, "price_max": 105.0,
    },
    "Backpack": {
        "base_min": 22.0, "base_max": 30.0, "base_avg": 26.0,
        "ship_min": 8.0,  "ship_max": 14.0, "ship_avg": 11.0,
        "margin_target": 0.60,
        "price_min": 75.0, "price_target": 85.0, "price_max": 95.0,
    },
    "Swim Trunk": {
        "base_min": 14.0, "base_max": 20.0, "base_avg": 17.0,
        "ship_min": 6.0,  "ship_max": 10.0, "ship_avg": 8.0,
        "margin_target": 0.58,
        "price_min": 45.0, "price_target": 55.0, "price_max": 65.0,
    },
    "Puffer": {
        "base_min": 38.0, "base_max": 55.0, "base_avg": 46.0,
        "ship_min": 8.0,  "ship_max": 14.0, "ship_avg": 11.0,
        "margin_target": 0.58,
        "price_min": 109.0, "price_target": 125.0, "price_max": 139.0,
    },
    "Windbreaker": {
        "base_min": 28.0, "base_max": 38.0, "base_avg": 33.0,
        "ship_min": 8.0,  "ship_max": 12.0, "ship_avg": 10.0,
        "margin_target": 0.60,
        "price_min": 95.0, "price_target": 109.0, "price_max": 125.0,
    },
    "Travel Bag": {
        "base_min": 30.0, "base_max": 42.0, "base_avg": 36.0,
        "ship_min": 8.0,  "ship_max": 14.0, "ship_avg": 11.0,
        "margin_target": 0.60,
        "price_min": 85.0, "price_target": 99.0, "price_max": 115.0,
    },
    "Duffle": {
        "base_min": 30.0, "base_max": 42.0, "base_avg": 36.0,
        "ship_min": 8.0,  "ship_max": 14.0, "ship_avg": 11.0,
        "margin_target": 0.60,
        "price_min": 85.0, "price_target": 99.0, "price_max": 115.0,
    },
    "One-Piece": {
        "base_min": 18.0, "base_max": 24.0, "base_avg": 21.0,
        "ship_min": 6.0,  "ship_max": 10.0, "ship_avg": 8.0,
        "margin_target": 0.58,
        "price_min": 55.0, "price_target": 65.0, "price_max": 75.0,
    },
    "Hawaiian": {
        "base_min": 20.0, "base_max": 28.0, "base_avg": 24.0,
        "ship_min": 6.0,  "ship_max": 10.0, "ship_avg": 8.0,
        "margin_target": 0.58,
        "price_min": 65.0, "price_target": 75.0, "price_max": 85.0,
    },
    "Lounge Pant": {
        "base_min": 16.0, "base_max": 22.0, "base_avg": 19.0,
        "ship_min": 6.0,  "ship_max": 10.0, "ship_avg": 8.0,
        "margin_target": 0.58,
        "price_min": 55.0, "price_target": 65.0, "price_max": 75.0,
    },
    "Hoodie": {
        "base_min": 24.0, "base_max": 32.0, "base_avg": 28.0,
        "ship_min": 6.0,  "ship_max": 10.0, "ship_avg": 8.0,
        "margin_target": 0.58,
        "price_min": 65.0, "price_target": 75.0, "price_max": 85.0,
    },
    "Dress": {
        "base_min": 20.0, "base_max": 28.0, "base_avg": 24.0,
        "ship_min": 6.0,  "ship_max": 10.0, "ship_avg": 8.0,
        "margin_target": 0.58,
        "price_min": 55.0, "price_target": 65.0, "price_max": 75.0,
    },
    "Slipper": {
        "base_min": 12.0, "base_max": 18.0, "base_avg": 15.0,
        "ship_min": 4.0,  "ship_max": 7.0,  "ship_avg": 5.5,
        "margin_target": 0.55,
        "price_min": 35.0, "price_target": 45.0, "price_max": 55.0,
    },
    "Tee": {
        "base_min": 16.0, "base_max": 22.0, "base_avg": 19.0,
        "ship_min": 4.0,  "ship_max": 7.0,  "ship_avg": 5.5,
        "margin_target": 0.55,
        "price_min": 45.0, "price_target": 55.0, "price_max": 65.0,
    },
    "Short": {
        "base_min": 14.0, "base_max": 20.0, "base_avg": 17.0,
        "ship_min": 4.0,  "ship_max": 7.0,  "ship_avg": 5.5,
        "margin_target": 0.55,
        "price_min": 45.0, "price_target": 55.0, "price_max": 65.0,
    },
    "Flip Flop": {
        "base_min": 10.0, "base_max": 16.0, "base_avg": 13.0,
        "ship_min": 4.0,  "ship_max": 7.0,  "ship_avg": 5.5,
        "margin_target": 0.55,
        "price_min": 35.0, "price_target": 45.0, "price_max": 55.0,
    },
}

# Formati prezzo approvati BKS (round o .95)
APPROVED_PRICES = [
    35, 39, 45, 49, 55, 59, 65, 69, 75, 79,
    85, 89, 95, 99, 105, 109, 115, 119, 125, 129, 135, 139
]

# ─────────────────────────────────────────────────────────────────────────────
# 2. FUNZIONI CORE
# ─────────────────────────────────────────────────────────────────────────────

def detect_category(title: str, product_type: str) -> str:
    """Rileva la categoria dal titolo prodotto o type."""
    text = (title + " " + (product_type or "")).lower()
    
    mapping = {
        "Sneaker":      ["sneaker"],
        "Backpack":     ["backpack"],
        "Swim Trunk":   ["swim trunk", "swim trunks"],
        "Puffer":       ["puffer"],
        "Windbreaker":  ["windbreaker"],
        "Travel Bag":   ["travel bag", "duffle", "duffel"],
        "Duffle":       ["duffle", "duffel"],
        "One-Piece":    ["one-piece", "one piece", "swimsuit"],
        "Hawaiian":     ["hawaiian"],
        "Lounge Pant":  ["lounge pant", "lounge pants"],
        "Hoodie":       ["hoodie", "pullover"],
        "Dress":        ["dress", "racerback"],
        "Slipper":      ["slipper"],
        "Tee":          ["tee", "t-shirt", "cut & sew"],
        "Short":        ["short", "shorts"],
        "Flip Flop":    ["flip flop"],
    }
    
    for category, keywords in mapping.items():
        for kw in keywords:
            if kw in text:
                return category
    return "Unknown"


def get_cost_data(category: str) -> dict:
    """Ritorna i dati di costo per una categoria."""
    for key in COST_TABLE:
        if key.lower() in category.lower() or category.lower() in key.lower():
            return COST_TABLE[key]
    return None


def calc_margin(price: float, cost_total: float) -> float:
    """Calcola margine % su prezzo vendita."""
    if price <= 0:
        return 0
    return (price - cost_total) / price * 100


def suggest_price(cost_total: float, margin_target: float,
                  price_min: float, price_max: float) -> float:
    """Suggerisce il prezzo ottimale dal formato approvato BKS."""
    min_price = cost_total / (1 - margin_target)
    
    # Trova il primo prezzo approvato che soddisfa il margine target
    candidates = [p for p in APPROVED_PRICES if p >= min_price]
    if not candidates:
        # Nessun prezzo approvato sufficiente — usa il calcolo diretto arrotondato
        raw = cost_total / (1 - margin_target)
        return round(raw / 5) * 5  # arrotonda al 5 più vicino
    
    suggested = candidates[0]
    
    # Preferisci il prezzo target di categoria se il margine regge
    if price_min <= suggested <= price_max:
        return suggested
    return suggested


def traffic_light(margin: float, price: float,
                  price_min: float, price_max: float,
                  margin_target: float) -> tuple[str, str]:
    """
    Ritorna (semaforo, motivo).
    🟢 OK · 🟡 ALERT · 🔴 STOP
    """
    if margin < 40:
        return "🔴 STOP", f"Margine {margin:.1f}% — sotto il minimo assoluto 40%"
    if margin < 45:
        return "🔴 STOP", f"Margine {margin:.1f}% — sotto il minimo 45%"
    if price < price_min:
        return "🟡 ALERT", f"Prezzo €{price} sotto il minimo di categoria €{price_min}"
    if price > price_max:
        return "🟡 ALERT", f"Prezzo €{price} sopra il massimo di categoria €{price_max}"
    if margin < margin_target * 100 - 5:
        return "🟡 ALERT", f"Margine {margin:.1f}% — sotto target {margin_target*100:.0f}%"
    if price not in APPROVED_PRICES:
        return "🟡 ALERT", f"Formato prezzo €{price} non nel listino approvato BKS"
    return "🟢 OK", f"Margine {margin:.1f}% — prezzo nel range corretto"


# ─────────────────────────────────────────────────────────────────────────────
# 3. LETTURA CSV SHOPIFY
# ─────────────────────────────────────────────────────────────────────────────

def load_shopify_csv(filepath: str) -> pd.DataFrame:
    """Carica e deduplicazione CSV Shopify export."""
    df = pd.read_csv(filepath, dtype=str)
    
    # Colonne chiave
    required = ["Title", "Variant Price", "Type"]
    for col in required:
        if col not in df.columns:
            # Prova alias comuni
            aliases = {
                "Title": ["title", "Name", "name"],
                "Variant Price": ["Price", "price", "variant_price"],
                "Type": ["Product Type", "product_type", "type"],
            }
            found = False
            for alias in aliases.get(col, []):
                if alias in df.columns:
                    df[col] = df[alias]
                    found = True
                    break
            if not found:
                print(f"⚠️  Colonna '{col}' non trovata — usando stringa vuota")
                df[col] = ""
    
    # Deduplica: una riga per prodotto (prima occorrenza = header row con Title)
    products = df[df["Title"].notna() & (df["Title"].str.strip() != "")].copy()
    products = products.drop_duplicates(subset=["Title"], keep="first")
    
    return products


# ─────────────────────────────────────────────────────────────────────────────
# 4. LETTURA COSTI PRINTIFY (da file opzionale)
# ─────────────────────────────────────────────────────────────────────────────

def load_costs_csv(filepath: str) -> dict:
    """
    Carica tabella costi custom da CSV.
    Formato atteso: Title,BaseCost,ShippingEU
    """
    if not filepath or not Path(filepath).exists():
        return {}
    
    df = pd.read_csv(filepath, dtype=str)
    costs = {}
    for _, row in df.iterrows():
        title = str(row.get("Title", "")).strip()
        if title:
            costs[title] = {
                "base": float(str(row.get("BaseCost", 0)).replace(",", ".")),
                "ship": float(str(row.get("ShippingEU", 0)).replace(",", ".")),
            }
    return costs


# ─────────────────────────────────────────────────────────────────────────────
# 5. AUDIT PRINCIPALE
# ─────────────────────────────────────────────────────────────────────────────

def run_audit(shopify_csv: str, costs_csv: str = None) -> tuple[pd.DataFrame, list]:
    """
    Esegue l'audit completo.
    Ritorna (DataFrame risultati, lista prodotti da correggere).
    """
    print(f"\n{'='*60}")
    print("BKS STUDIO — PRICE AUDIT")
    print(f"{'='*60}")
    print(f"CSV Shopify: {shopify_csv}")
    print(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
    
    # Carica dati
    products = load_shopify_csv(shopify_csv)
    custom_costs = load_costs_csv(costs_csv) if costs_csv else {}
    
    print(f"Prodotti trovati: {len(products)}\n")
    
    results = []
    corrections = []
    
    for _, row in products.iterrows():
        title = str(row.get("Title", "")).strip()
        price_str = str(row.get("Variant Price", "0")).replace(",", ".").replace("€", "").strip()
        product_type = str(row.get("Type", "")).strip()
        handle = str(row.get("Handle", "")).strip()
        
        # Prezzo attuale
        try:
            current_price = float(price_str)
        except ValueError:
            current_price = 0.0
        
        # Categoria
        category = detect_category(title, product_type)
        cost_data = get_cost_data(category)
        
        if not cost_data:
            results.append({
                "Handle": handle,
                "Title": title,
                "Category": "UNKNOWN",
                "Current Price (€)": current_price,
                "Base Cost (€)": "?",
                "Shipping EU (€)": "?",
                "Total Cost (€)": "?",
                "Margin %": "?",
                "Status": "⚪ SKIP",
                "Note": "Categoria non riconosciuta — verificare manualmente",
                "Suggested Price (€)": "?",
                "Delta (€)": "?",
            })
            continue
        
        # Usa costi custom se disponibili, altrimenti usa la media della tabella
        if title in custom_costs:
            base_cost = custom_costs[title]["base"]
            ship_cost = custom_costs[title]["ship"]
        else:
            base_cost = cost_data["base_avg"]
            ship_cost = cost_data["ship_avg"]
        
        total_cost = base_cost + ship_cost
        margin = calc_margin(current_price, total_cost)
        status, note = traffic_light(
            margin, current_price,
            cost_data["price_min"], cost_data["price_max"],
            cost_data["margin_target"]
        )
        
        # Prezzo suggerito
        suggested = suggest_price(
            total_cost,
            cost_data["margin_target"],
            cost_data["price_min"],
            cost_data["price_max"]
        )
        suggested_margin = calc_margin(suggested, total_cost)
        delta = suggested - current_price
        
        result_row = {
            "Handle": handle,
            "Title": title,
            "Category": category,
            "Current Price (€)": round(current_price, 2),
            "Base Cost (€)": round(base_cost, 2),
            "Shipping EU (€)": round(ship_cost, 2),
            "Total Cost (€)": round(total_cost, 2),
            "Margin %": round(margin, 1),
            "Margin Target %": int(cost_data["margin_target"] * 100),
            "Status": status,
            "Note": note,
            "Suggested Price (€)": round(suggested, 2),
            "Suggested Margin %": round(suggested_margin, 1),
            "Delta (€)": round(delta, 2),
        }
        results.append(result_row)
        
        # Aggiungi alla lista correzioni se non OK
        if status != "🟢 OK" and delta != 0:
            corrections.append({
                "handle": handle,
                "title": title,
                "current_price": current_price,
                "suggested_price": suggested,
                "delta": delta,
                "status": status,
                "margin_current": round(margin, 1),
                "margin_suggested": round(suggested_margin, 1),
            })
    
    df_results = pd.DataFrame(results)
    
    # Stampa summary
    total = len(df_results)
    ok = len(df_results[df_results["Status"] == "🟢 OK"])
    alert = len(df_results[df_results["Status"].str.startswith("🟡")])
    stop = len(df_results[df_results["Status"].str.startswith("🔴")])
    skip = len(df_results[df_results["Status"] == "⚪ SKIP"])
    
    print(f"{'─'*60}")
    print(f"RISULTATI AUDIT")
    print(f"{'─'*60}")
    print(f"  🟢 OK:      {ok:3d} prodotti ({ok/total*100:.1f}%)")
    print(f"  🟡 ALERT:   {alert:3d} prodotti ({alert/total*100:.1f}%)")
    print(f"  🔴 STOP:    {stop:3d} prodotti ({stop/total*100:.1f}%)")
    print(f"  ⚪ SKIP:    {skip:3d} prodotti (categoria non riconosciuta)")
    print(f"  TOTALE:     {total:3d} prodotti")
    print(f"{'─'*60}")
    
    if corrections:
        avg_delta = sum(c["delta"] for c in corrections) / len(corrections)
        print(f"\n  Prodotti che richiedono correzione prezzo: {len(corrections)}")
        print(f"  Delta medio: €{avg_delta:+.2f}")
        
        # Top 5 urgenti
        urgent = sorted(corrections, key=lambda x: x["margin_current"])[:5]
        print(f"\n  TOP 5 URGENTI (margine più basso):")
        for c in urgent:
            print(f"  → {c['title'][:45]:<45} {c['margin_current']:5.1f}% → "
                  f"€{c['current_price']:.0f} → €{c['suggested_price']:.0f}")
    
    print()
    return df_results, corrections


# ─────────────────────────────────────────────────────────────────────────────
# 6. EXPORT REPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_report(df: pd.DataFrame, corrections: list, output_dir: str = "."):
    """Esporta report Excel + CSV correzioni."""
    out = Path(output_dir)
    out.mkdir(exist_ok=True)
    
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    
    # ── Excel con colori semaforo ──────────────────────────────────────────
    excel_path = out / f"bks_price_audit_{ts}.xlsx"
    
    try:
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Audit", index=False)
            ws = writer.sheets["Audit"]
            
            # Larghezza colonne
            for i, col in enumerate(df.columns, 1):
                ws.column_dimensions[get_column_letter(i)].width = max(
                    len(str(col)) + 2, 14
                )
            
            # Colori per Status
            color_map = {
                "🟢": "C6EFCE",  # verde
                "🟡": "FFEB9C",  # giallo
                "🔴": "FFC7CE",  # rosso
                "⚪": "EEEEEE",  # grigio
            }
            
            status_col = df.columns.get_loc("Status") + 1
            for row_idx, status in enumerate(df["Status"], 2):
                for emoji, color in color_map.items():
                    if str(status).startswith(emoji):
                        fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )
                        for col_idx in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                        break
            
            # Header bold
            header_font = Font(bold=True)
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=1, column=col_idx).font = header_font
        
        print(f"✅ Report Excel:     {excel_path}")
    
    except ImportError:
        # Fallback CSV se openpyxl non disponibile
        csv_path = out / f"bks_price_audit_{ts}.csv"
        df.to_csv(csv_path, index=False)
        print(f"✅ Report CSV:       {csv_path}")
        excel_path = csv_path
    
    # ── CSV correzioni per import Shopify ─────────────────────────────────
    if corrections:
        corr_df = pd.DataFrame([{
            "Handle": c["handle"],
            "Title": c["title"],
            "Variant Price": c["suggested_price"],  # colonna Shopify
        } for c in corrections])
        
        corr_csv = out / f"bks_price_corrections_{ts}.csv"
        corr_df.to_csv(corr_csv, index=False)
        print(f"✅ CSV correzioni:   {corr_csv}")
        print(f"   → {len(corrections)} prodotti · importa in Shopify Admin > Products > Import")
    
    # ── JSON per API update ───────────────────────────────────────────────
    api_json = out / f"bks_price_corrections_api_{ts}.json"
    with open(api_json, "w", encoding="utf-8") as f:
        json.dump(corrections, f, indent=2, ensure_ascii=False)
    print(f"✅ JSON API:         {api_json}")
    
    return excel_path


# ─────────────────────────────────────────────────────────────────────────────
# 7. AGGIORNAMENTO VIA SHOPIFY API (opzionale)
# ─────────────────────────────────────────────────────────────────────────────

def update_prices_via_api(corrections: list, dry_run: bool = True):
    """
    Aggiorna i prezzi via Shopify Admin API.
    Richiede SHOPIFY_STORE e SHOPIFY_TOKEN nel .env
    
    dry_run=True → stampa solo le chiamate senza eseguirle
    """
    store = os.environ.get("SHOPIFY_STORE", "")
    token = os.environ.get("SHOPIFY_TOKEN", "")
    
    if not store or not token:
        print("\n⚠️  SHOPIFY_STORE e SHOPIFY_TOKEN non trovati nel .env")
        print("   Aggiornamento via API saltato — usa il CSV import manuale")
        return
    
    base_url = f"https://{store}.myshopify.com/admin/api/2024-01"
    headers = {
        "X-Shopify-Access-Token": token,
        "Content-Type": "application/json"
    }
    
    print(f"\n{'─'*60}")
    if dry_run:
        print("DRY RUN — Chiamate API che verrebbero eseguite:")
    else:
        print("AGGIORNAMENTO PREZZI VIA API:")
    print(f"{'─'*60}")
    
    success = 0
    errors = 0
    
    for c in corrections:
        handle = c["handle"]
        new_price = c["suggested_price"]
        
        if dry_run:
            print(f"  PUT /products/[id]/variants/[vid].json  "
                  f"price={new_price}  handle={handle}")
            continue
        
        # Step 1: trova il prodotto per handle
        r = requests.get(
            f"{base_url}/products.json",
            headers=headers,
            params={"handle": handle, "fields": "id,title,variants"}
        )
        
        if r.status_code != 200:
            print(f"  ❌ {handle} — errore ricerca: {r.status_code}")
            errors += 1
            continue
        
        data = r.json()
        products = data.get("products", [])
        
        if not products:
            print(f"  ⚠️  {handle} — prodotto non trovato")
            errors += 1
            continue
        
        product = products[0]
        product_id = product["id"]
        
        # Step 2: aggiorna il prezzo su tutte le varianti
        updated_vars = 0
        for variant in product.get("variants", []):
            vid = variant["id"]
            payload = {"variant": {"id": vid, "price": str(new_price)}}
            
            r2 = requests.put(
                f"{base_url}/variants/{vid}.json",
                headers=headers,
                json=payload
            )
            
            if r2.status_code == 200:
                updated_vars += 1
            else:
                print(f"    ⚠️  Variante {vid} — errore: {r2.status_code}")
        
        if updated_vars > 0:
            print(f"  ✅ {handle} — {updated_vars} varianti → €{new_price}")
            success += 1
        else:
            errors += 1
        
        # Rate limit Shopify: max 2 req/sec
        import time
        time.sleep(0.6)
    
    if not dry_run:
        print(f"\n  Aggiornati: {success} · Errori: {errors}")


# ─────────────────────────────────────────────────────────────────────────────
# 8. MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="BKS Studio — Price Audit & Update Tool"
    )
    parser.add_argument(
        "--csv", required=True,
        help="Percorso al CSV export Shopify (Products > Export)"
    )
    parser.add_argument(
        "--costs",
        help="Percorso CSV costi Printify custom (opzionale). "
             "Formato: Title,BaseCost,ShippingEU"
    )
    parser.add_argument(
        "--output", default=".",
        help="Cartella output per report (default: cartella corrente)"
    )
    parser.add_argument(
        "--update", action="store_true",
        help="Aggiorna i prezzi via Shopify API (richiede SHOPIFY_STORE e SHOPIFY_TOKEN nel .env)"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Mostra le chiamate API senza eseguirle (usa con --update)"
    )
    
    args = parser.parse_args()
    
    # Audit
    df_results, corrections = run_audit(args.csv, args.costs)
    
    # Export
    export_report(df_results, corrections, args.output)
    
    # API update (opzionale)
    if args.update or args.dry_run:
        update_prices_via_api(corrections, dry_run=not args.update or args.dry_run)
    
    print(f"\n{'='*60}")
    print("PROSSIMI PASSI:")
    print("  1. Apri bks_price_audit_[ts].xlsx — verifica il semaforo")
    print("  2. Per ogni 🔴 STOP: aggiornare il prezzo in Shopify prima possibile")
    print("  3. Importa bks_price_corrections_[ts].csv in Shopify Admin > Products > Import")
    print("     OPPURE esegui: python3 bks_price_audit.py --csv [file] --update")
    print("  4. Aggiorna i costi reali Printify nel file costs_table.csv")
    print("     per avere calcoli precisi invece dei valori medi")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
